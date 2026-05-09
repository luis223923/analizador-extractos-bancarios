"""
Sistema de Control de Deuda Bancaria
Aplicación Streamlit para la gestión y control de deuda bancaria empresarial.
"""

import io
import warnings
from datetime import date, datetime, timedelta

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

warnings.filterwarnings("ignore")

# ─── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Sistema de Control de Deuda Bancaria",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CSS corporativo ──────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 1.8rem; font-weight: 700; color: #1a3a5c;
        border-bottom: 3px solid #1a3a5c; padding-bottom: 0.4rem; margin-bottom: 1rem;
    }
    .alert-critica  { background:#fde8e8; border-left:5px solid #c0392b; padding:0.6rem 1rem; border-radius:6px; margin:4px 0; }
    .alert-alta     { background:#fef3e2; border-left:5px solid #e67e22; padding:0.6rem 1rem; border-radius:6px; margin:4px 0; }
    .alert-media    { background:#fefbe2; border-left:5px solid #f1c40f; padding:0.6rem 1rem; border-radius:6px; margin:4px 0; }
    .alert-info     { background:#e8f4fd; border-left:5px solid #2980b9; padding:0.6rem 1rem; border-radius:6px; margin:4px 0; }
    div[data-testid="metric-container"] {
        background:#f0f4f8; border-radius:8px; padding:0.8rem; border-left:4px solid #1a3a5c;
    }
</style>
""", unsafe_allow_html=True)

# ─── Constantes de dominio ────────────────────────────────────────────────────
TIPOS_OPERACION = [
    "Linea de Credito", "Prestamo Directo", "Boleta de Garantia",
    "Refinanciamiento", "Renovacion", "Sobregiro", "Carta de Credito", "Otra Facilidad",
]
ESTADOS_OPERACION = ["Vigente", "Pagado", "Vencido", "Renovado", "Refinanciado", "En Tramite", "Cancelado"]
ESTADOS_PAGO     = ["Pendiente", "Pagado", "Reprogramado", "Vencido", "Anulado"]
ESTADOS_BOLETA   = ["Vigente", "Vencida", "Liberada", "Renovada", "Ejecutada", "En Tramite"]
CLASIFICACIONES  = ["Vencido", "0 a 7 días", "8 a 30 días", "31 a 60 días", "61 a 90 días", "Más de 90 días"]
TIPOS_GARANTIA   = ["Hipoteca", "Prenda", "Fianza", "Boleta de Garantia", "Warrant", "Sin Garantia", "Otra"]
MONEDAS          = ["CLP", "USD", "EUR", "UF"]

HOY = date.today()


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════

def clasificar_vencimiento(dias) -> str:
    if pd.isna(dias):
        return "Sin fecha"
    dias = float(dias)
    if dias < 0:
        return "Vencido"
    elif dias <= 7:
        return "0 a 7 días"
    elif dias <= 30:
        return "8 a 30 días"
    elif dias <= 60:
        return "31 a 60 días"
    elif dias <= 90:
        return "61 a 90 días"
    else:
        return "Más de 90 días"


def normalizar_datos(dfs: dict) -> dict:
    """Asegura columnas obligatorias y calcula campos derivados."""
    today = pd.Timestamp(HOY)

    # ── Base_Deuda ──
    if "Base_Deuda" in dfs and not dfs["Base_Deuda"].empty:
        df = dfs["Base_Deuda"].copy()
        for col in ["Fecha_Desembolso", "Fecha_Vencimiento"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        for col in ["Tasa_Anual", "Saldo_Actual", "Monto_Original"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        if "Fecha_Vencimiento" in df.columns:
            df["Dias_Al_Vencimiento"] = (df["Fecha_Vencimiento"] - today).dt.days
            df["Clasificacion_Vencimiento"] = df["Dias_Al_Vencimiento"].apply(clasificar_vencimiento)
        if "Saldo_Actual" in df.columns and "Tasa_Anual" in df.columns:
            df["Interes_Estimado_Mensual"] = (df["Saldo_Actual"] * df["Tasa_Anual"] / 100 / 12).round(2)
        dfs["Base_Deuda"] = df

    # ── Boletas_Garantia ──
    if "Boletas_Garantia" in dfs and not dfs["Boletas_Garantia"].empty:
        df = dfs["Boletas_Garantia"].copy()
        if "Fecha_Vencimiento" in df.columns:
            df["Fecha_Vencimiento"] = pd.to_datetime(df["Fecha_Vencimiento"], errors="coerce")
            df["Dias_Al_Vencimiento"] = (df["Fecha_Vencimiento"] - today).dt.days
        if "Monto" in df.columns:
            df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce").fillna(0)
        dfs["Boletas_Garantia"] = df

    # ── Lineas_Credito ──
    if "Lineas_Credito" in dfs and not dfs["Lineas_Credito"].empty:
        df = dfs["Lineas_Credito"].copy()
        for col in ["Monto_Aprobado", "Monto_Utilizado"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        if "Monto_Aprobado" in df.columns and "Monto_Utilizado" in df.columns:
            df["Monto_Disponible"] = df["Monto_Aprobado"] - df["Monto_Utilizado"]
        if "Fecha_Vencimiento" in df.columns:
            df["Fecha_Vencimiento"] = pd.to_datetime(df["Fecha_Vencimiento"], errors="coerce")
        dfs["Lineas_Credito"] = df

    # ── Cronograma_Pagos ──
    if "Cronograma_Pagos" in dfs and not dfs["Cronograma_Pagos"].empty:
        df = dfs["Cronograma_Pagos"].copy()
        if "Fecha_Pago" in df.columns:
            df["Fecha_Pago"] = pd.to_datetime(df["Fecha_Pago"], errors="coerce")
        for col in ["Capital", "Interes", "Comision", "Seguro", "Otros_Cargos", "Total_Pagar"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        dfs["Cronograma_Pagos"] = df

    # ── Pagos_Realizados ──
    if "Pagos_Realizados" in dfs and not dfs["Pagos_Realizados"].empty:
        df = dfs["Pagos_Realizados"].copy()
        if "Fecha_Pago" in df.columns:
            df["Fecha_Pago"] = pd.to_datetime(df["Fecha_Pago"], errors="coerce")
        for col in ["Capital_Pagado", "Interes_Pagado", "Comision_Pagada", "Total_Pagado"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        dfs["Pagos_Realizados"] = df

    return dfs


def calcular_kpis(dfs: dict) -> dict:
    """Calcula todos los KPIs del sistema."""
    kpis = {}
    today = pd.Timestamp(HOY)

    df_deuda  = dfs.get("Base_Deuda", pd.DataFrame())
    df_lineas = dfs.get("Lineas_Credito", pd.DataFrame())
    df_boletas= dfs.get("Boletas_Garantia", pd.DataFrame())
    df_cron   = dfs.get("Cronograma_Pagos", pd.DataFrame())

    # ── Deuda bancaria ──
    if not df_deuda.empty and "Saldo_Actual" in df_deuda.columns:
        kpis["deuda_total"]      = df_deuda["Saldo_Actual"].sum()
        kpis["deuda_por_banco"]  = df_deuda.groupby("Banco")["Saldo_Actual"].sum().reset_index() if "Banco" in df_deuda.columns else pd.DataFrame()
        kpis["deuda_por_empresa"]= df_deuda.groupby("Empresa")["Saldo_Actual"].sum().reset_index() if "Empresa" in df_deuda.columns else pd.DataFrame()
        kpis["deuda_por_moneda"] = df_deuda.groupby("Moneda")["Saldo_Actual"].sum().reset_index() if "Moneda" in df_deuda.columns else pd.DataFrame()
        kpis["deuda_por_tipo"]   = df_deuda.groupby("Tipo_Operacion")["Saldo_Actual"].sum().reset_index() if "Tipo_Operacion" in df_deuda.columns else pd.DataFrame()

        saldo_total = df_deuda["Saldo_Actual"].sum()
        if saldo_total > 0 and "Tasa_Anual" in df_deuda.columns:
            kpis["tasa_promedio"] = (df_deuda["Saldo_Actual"] * df_deuda["Tasa_Anual"]).sum() / saldo_total
        else:
            kpis["tasa_promedio"] = 0.0

        if "Dias_Al_Vencimiento" in df_deuda.columns and "Estado" in df_deuda.columns:
            vig = df_deuda["Estado"] == "Vigente"
            dias = df_deuda["Dias_Al_Vencimiento"]
            kpis["ops_vencidas"]     = int((vig & (dias < 0)).sum())
            kpis["ops_por_vencer_7"] = int((vig & (dias >= 0) & (dias <= 7)).sum())
            kpis["ops_por_vencer_30"]= int((vig & (dias >= 0) & (dias <= 30)).sum())
        else:
            kpis["ops_vencidas"] = kpis["ops_por_vencer_7"] = kpis["ops_por_vencer_30"] = 0
    else:
        kpis.update({
            "deuda_total": 0, "deuda_por_banco": pd.DataFrame(), "deuda_por_empresa": pd.DataFrame(),
            "deuda_por_moneda": pd.DataFrame(), "deuda_por_tipo": pd.DataFrame(),
            "tasa_promedio": 0.0, "ops_vencidas": 0, "ops_por_vencer_7": 0, "ops_por_vencer_30": 0,
        })

    # ── Cronograma ──
    if not df_cron.empty and "Fecha_Pago" in df_cron.columns:
        for dias, llave in [(7, "7d"), (30, "30d"), (60, "60d"), (90, "90d")]:
            mask = (df_cron["Fecha_Pago"] >= today) & (df_cron["Fecha_Pago"] <= today + pd.Timedelta(days=dias))
            sub = df_cron[mask]
            kpis[f"capital_{llave}"] = sub["Capital"].sum() if "Capital" in sub.columns else 0
            kpis[f"interes_{llave}"] = sub["Interes"].sum() if "Interes" in sub.columns else 0
            kpis[f"total_{llave}"]   = sub["Total_Pagar"].sum() if "Total_Pagar" in sub.columns else 0
    else:
        for llave in ["7d", "30d", "60d", "90d"]:
            kpis[f"capital_{llave}"] = kpis[f"interes_{llave}"] = kpis[f"total_{llave}"] = 0

    # ── Líneas de crédito ──
    if not df_lineas.empty:
        kpis["lineas_aprobadas"]  = df_lineas["Monto_Aprobado"].sum()  if "Monto_Aprobado"  in df_lineas.columns else 0
        kpis["lineas_utilizadas"] = df_lineas["Monto_Utilizado"].sum() if "Monto_Utilizado" in df_lineas.columns else 0
        kpis["lineas_disponibles"]= df_lineas["Monto_Disponible"].sum()if "Monto_Disponible"in df_lineas.columns else 0
    else:
        kpis["lineas_aprobadas"] = kpis["lineas_utilizadas"] = kpis["lineas_disponibles"] = 0

    # ── Boletas ──
    if not df_boletas.empty and "Estado" in df_boletas.columns:
        kpis["boletas_vigentes"]    = int((df_boletas["Estado"] == "Vigente").sum())
        kpis["boletas_vencidas"]    = int((df_boletas["Estado"] == "Vencida").sum())
        if "Dias_Al_Vencimiento" in df_boletas.columns:
            vig_b = df_boletas["Estado"] == "Vigente"
            dias_b= df_boletas["Dias_Al_Vencimiento"]
            kpis["boletas_por_vencer_30"] = int((vig_b & (dias_b >= 0) & (dias_b <= 30)).sum())
        else:
            kpis["boletas_por_vencer_30"] = 0
        kpis["monto_boletas_vigentes"] = df_boletas.loc[df_boletas["Estado"] == "Vigente", "Monto"].sum() if "Monto" in df_boletas.columns else 0
    else:
        kpis["boletas_vigentes"] = kpis["boletas_vencidas"] = kpis["boletas_por_vencer_30"] = 0
        kpis["monto_boletas_vigentes"] = 0

    kpis["exposicion_total"] = kpis["deuda_total"] + kpis["lineas_utilizadas"] + kpis["monto_boletas_vigentes"]
    return kpis


def calcular_alertas(dfs: dict) -> pd.DataFrame:
    """Genera tabla de alertas con nivel, mensaje y referencia."""
    alertas = []
    today = pd.Timestamp(HOY)

    def add(nivel: str, mensaje: str, ref: str = ""):
        alertas.append({"Nivel": nivel, "Mensaje": mensaje, "Referencia": str(ref)})

    df_deuda  = dfs.get("Base_Deuda", pd.DataFrame())
    df_boletas= dfs.get("Boletas_Garantia", pd.DataFrame())
    df_lineas = dfs.get("Lineas_Credito", pd.DataFrame())
    df_cron   = dfs.get("Cronograma_Pagos", pd.DataFrame())
    df_pagos  = dfs.get("Pagos_Realizados", pd.DataFrame())

    # ── Validaciones Base_Deuda ──
    if not df_deuda.empty:
        cols_dup = [c for c in ["Banco", "Numero_Operacion_Banco", "Monto_Original", "Fecha_Desembolso"] if c in df_deuda.columns]
        ids_dup = set()
        if len(cols_dup) == 4:
            dup_mask = df_deuda.duplicated(subset=cols_dup, keep=False)
            ids_dup = set(df_deuda[dup_mask].get("ID_Operacion", pd.Series()).astype(str).tolist())

        ids_con_cron = set()
        if not df_cron.empty and "ID_Operacion" in df_cron.columns:
            ids_con_cron = set(df_cron["ID_Operacion"].dropna().astype(str))

        for _, row in df_deuda.iterrows():
            op_id = str(row.get("ID_Operacion", ""))
            num   = str(row.get("Numero_Operacion_Banco", ""))
            ref   = f"{op_id} / {num}"

            # Campos obligatorios
            for campo, etiq in [("Empresa","Empresa"), ("Banco","Banco"), ("Moneda","Moneda")]:
                if pd.isna(row.get(campo)) or str(row.get(campo,"")).strip() == "":
                    add("Media", f"Operación sin {etiq} registrada", ref)
            if pd.isna(row.get("Numero_Operacion_Banco")) or str(row.get("Numero_Operacion_Banco","")).strip() == "":
                add("Media", "Operación sin número de operación bancaria", ref)

            # Saldo > monto
            saldo = float(row.get("Saldo_Actual", 0) or 0)
            monto = float(row.get("Monto_Original", 0) or 0)
            if monto > 0 and saldo > monto:
                add("Alta", f"Saldo_Actual ({saldo:,.0f}) supera Monto_Original ({monto:,.0f})", ref)

            estado = str(row.get("Estado", ""))
            dias   = row.get("Dias_Al_Vencimiento", None)
            tipo   = str(row.get("Tipo_Operacion", ""))

            # Vencimientos
            if pd.notna(dias):
                dias_f = float(dias)
                if dias_f < 0 and estado == "Vigente":
                    add("Crítica", "Operación VENCIDA con estado Vigente", ref)
                elif 0 <= dias_f <= 7 and estado == "Vigente":
                    add("Alta", f"Operación vence en {int(dias_f)} días", ref)
                elif 8 <= dias_f <= 30 and estado == "Vigente":
                    add("Media", f"Operación vence en {int(dias_f)} días", ref)
                # Vencida sin estado de cierre
                if dias_f < 0 and estado not in ["Pagado", "Renovado", "Refinanciado", "Cancelado", "Vencido"]:
                    add("Crítica", "Operación vencida sin estado de cierre correcto", ref)

            # Tasa faltante en créditos
            if tipo in ["Prestamo Directo", "Linea de Credito", "Refinanciamiento", "Renovacion"]:
                tasa = row.get("Tasa_Anual", None)
                if pd.isna(tasa) or float(tasa or 0) == 0:
                    add("Media", "Operación sin Tasa_Anual registrada", ref)

            # Refinanciado sin nueva operación
            if estado == "Refinanciado":
                nueva = row.get("ID_Nueva_Operacion", None)
                if pd.isna(nueva) or str(nueva).strip() == "":
                    add("Alta", "Operación Refinanciada sin ID_Nueva_Operacion", ref)

            # Sin garantía
            gar = row.get("Garantia_Asociada", None)
            if pd.isna(gar) or str(gar).strip() == "":
                add("Informativa", "Operación sin garantía asociada", ref)

            # Sin observaciones
            obs = row.get("Observaciones", None)
            if pd.isna(obs) or str(obs).strip() == "":
                add("Informativa", "Operación sin observaciones", ref)

            # Duplicado
            if op_id in ids_dup:
                add("Alta", "Posible operación duplicada (Banco+Número+Monto+Fecha)", ref)

            # Sin cronograma
            if estado == "Vigente" and op_id not in ids_con_cron and ids_con_cron:
                add("Informativa", "Operación vigente sin cronograma de pagos", ref)

    # ── Boletas ──
    if not df_boletas.empty:
        for _, row in df_boletas.iterrows():
            ref   = f"{row.get('ID_Boleta','')} / {row.get('Numero_Boleta','')}"
            dias  = row.get("Dias_Al_Vencimiento", None)
            estado= str(row.get("Estado", ""))
            if pd.notna(dias):
                dias_f = float(dias)
                if dias_f < 0 and estado == "Vigente":
                    add("Crítica", "Boleta de Garantía VENCIDA con estado Vigente", ref)
                elif 0 <= dias_f <= 7 and estado == "Vigente":
                    add("Alta", f"Boleta vence en {int(dias_f)} días", ref)
                elif 8 <= dias_f <= 30 and estado == "Vigente":
                    add("Media", f"Boleta vence en {int(dias_f)} días", ref)

    # ── Líneas de crédito ──
    if not df_lineas.empty:
        for _, row in df_lineas.iterrows():
            ref   = f"{row.get('ID_Linea','')} / {row.get('Numero_Linea','')}"
            estado= str(row.get("Estado", ""))
            fv    = row.get("Fecha_Vencimiento", None)
            if pd.notna(fv):
                dias_f = (pd.Timestamp(fv) - today).days
                if dias_f < 0 and estado == "Vigente":
                    add("Crítica", "Línea de Crédito VENCIDA con estado Vigente", ref)
                elif 0 <= dias_f <= 7:
                    add("Alta", f"Línea de Crédito vence en {int(dias_f)} días", ref)
                elif 8 <= dias_f <= 30:
                    add("Media", f"Línea de Crédito vence en {int(dias_f)} días", ref)

    # ── Pagos realizados ──
    if not df_pagos.empty:
        for _, row in df_pagos.iterrows():
            ref = str(row.get("ID_Pago_Realizado", ""))
            comp = row.get("Comprobante", None)
            if pd.isna(comp) or str(comp).strip() == "":
                add("Informativa", "Pago realizado sin número de comprobante", ref)

    if not alertas:
        return pd.DataFrame(columns=["Nivel", "Mensaje", "Referencia"])

    df_al = pd.DataFrame(alertas)
    orden = {"Crítica": 0, "Alta": 1, "Media": 2, "Informativa": 3}
    df_al["_ord"] = df_al["Nivel"].map(orden)
    return df_al.sort_values("_ord").drop(columns="_ord").reset_index(drop=True)


def calcular_trazabilidad(dfs: dict) -> pd.DataFrame:
    """Construye tabla de trazabilidad entre operaciones."""
    df_ref   = dfs.get("Refinanciamientos", pd.DataFrame())
    df_deuda = dfs.get("Base_Deuda", pd.DataFrame())

    # Usar hoja Refinanciamientos si existe
    if not df_ref.empty:
        result = df_ref.copy()
        if not df_deuda.empty and "ID_Operacion" in df_deuda.columns and "Numero_Operacion_Banco" in df_deuda.columns:
            mapa = dict(zip(df_deuda["ID_Operacion"].astype(str), df_deuda["Numero_Operacion_Banco"]))
            if "ID_Operacion_Anterior" in result.columns:
                result["Numero_Operacion_Anterior"] = result["ID_Operacion_Anterior"].astype(str).map(mapa)
            if "ID_Operacion_Nueva" in result.columns:
                result["Numero_Operacion_Nueva"] = result["ID_Operacion_Nueva"].astype(str).map(mapa)
        return result

    # Construir desde Base_Deuda si hay columna ID_Operacion_Reemplazada
    if not df_deuda.empty and "ID_Operacion_Reemplazada" in df_deuda.columns:
        mapa_num  = {}
        mapa_monto= {}
        if "Numero_Operacion_Banco" in df_deuda.columns:
            mapa_num = dict(zip(df_deuda["ID_Operacion"].astype(str), df_deuda["Numero_Operacion_Banco"]))
        if "Monto_Original" in df_deuda.columns:
            mapa_monto = dict(zip(df_deuda["ID_Operacion"].astype(str), df_deuda["Monto_Original"]))

        filas = []
        for _, row in df_deuda[df_deuda["ID_Operacion_Reemplazada"].notna()].iterrows():
            ant_id = str(row.get("ID_Operacion_Reemplazada", ""))
            filas.append({
                "ID_Operacion_Anterior":     ant_id,
                "Numero_Operacion_Anterior": mapa_num.get(ant_id, ""),
                "Banco":                     row.get("Banco", ""),
                "Empresa":                   row.get("Empresa", ""),
                "Monto_Cancelado":           mapa_monto.get(ant_id, 0),
                "ID_Operacion_Nueva":        row.get("ID_Operacion", ""),
                "Numero_Operacion_Nueva":    mapa_num.get(str(row.get("ID_Operacion", "")), ""),
                "Monto_Nuevo":               row.get("Monto_Original", 0),
                "Observaciones":             row.get("Observaciones", ""),
            })
        return pd.DataFrame(filas) if filas else pd.DataFrame()

    return pd.DataFrame()


def calcular_exposicion_por_banco(dfs: dict) -> pd.DataFrame:
    filas: dict = {}

    df_deuda  = dfs.get("Base_Deuda", pd.DataFrame())
    df_lineas = dfs.get("Lineas_Credito", pd.DataFrame())
    df_boletas= dfs.get("Boletas_Garantia", pd.DataFrame())

    if not df_deuda.empty and "Banco" in df_deuda.columns and "Saldo_Actual" in df_deuda.columns:
        for banco, saldo in df_deuda.groupby("Banco")["Saldo_Actual"].sum().items():
            filas.setdefault(banco, {"Banco": banco, "Deuda": 0, "Lineas_Utilizadas": 0, "Boletas": 0})
            filas[banco]["Deuda"] += saldo

    if not df_lineas.empty and "Banco" in df_lineas.columns and "Monto_Utilizado" in df_lineas.columns:
        for banco, util in df_lineas.groupby("Banco")["Monto_Utilizado"].sum().items():
            filas.setdefault(banco, {"Banco": banco, "Deuda": 0, "Lineas_Utilizadas": 0, "Boletas": 0})
            filas[banco]["Lineas_Utilizadas"] += util

    if not df_boletas.empty and "Banco" in df_boletas.columns and "Monto" in df_boletas.columns:
        vig = df_boletas[df_boletas.get("Estado", pd.Series(dtype=str)) == "Vigente"] if "Estado" in df_boletas.columns else df_boletas
        for banco, monto in vig.groupby("Banco")["Monto"].sum().items():
            filas.setdefault(banco, {"Banco": banco, "Deuda": 0, "Lineas_Utilizadas": 0, "Boletas": 0})
            filas[banco]["Boletas"] += monto

    if not filas:
        return pd.DataFrame()

    df = pd.DataFrame(filas.values())
    df["Exposicion_Total"] = df["Deuda"] + df["Lineas_Utilizadas"] + df["Boletas"]
    return df.sort_values("Exposicion_Total", ascending=False).reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PLANTILLA EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def crear_plantilla_excel(con_ejemplo: bool = False) -> bytes:
    """Genera plantilla Excel con todas las hojas del sistema."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", fgColor="1A3A5C")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_sheet(nombre: str, columnas: list, filas: list = None):
        ws = wb.create_sheet(nombre)
        ws.row_dimensions[1].height = 28
        for ci, col in enumerate(columnas, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = hdr_align; c.border = brd
        if filas:
            for ri, fila in enumerate(filas, 2):
                for ci, val in enumerate(fila, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = brd
                    c.alignment = Alignment(vertical="center")
        # Auto-ancho
        for ci, col_cells in enumerate(ws.columns, 1):
            length = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(length + 2, 12), 38)
        return ws

    # ── Base_Deuda ──
    cols_bd = [
        "ID_Operacion","Empresa","Banco","Tipo_Operacion","Numero_Operacion_Banco",
        "Moneda","Monto_Original","Saldo_Actual","Tasa_Anual","Fecha_Desembolso",
        "Fecha_Vencimiento","Plazo_Dias","Estado","Cuenta_Debito","Destino_Fondos",
        "Garantia_Asociada","ID_Operacion_Reemplazada","ID_Nueva_Operacion",
        "Observaciones","Dias_Al_Vencimiento","Interes_Estimado_Mensual","Clasificacion_Vencimiento",
    ]
    filas_bd = []
    if con_ejemplo:
        filas_bd = [
            ["OP001","Empresa A","Banco Chile","Prestamo Directo","42356789","CLP",500_000_000,450_000_000,6.5,date(2024,3,1),date(2025,9,1),549,"Vigente","001-001-001","Capital de Trabajo","GAR001","","","",None,None,None],
            ["OP002","Empresa A","Banco BCI","Prestamo Directo","78901234","CLP",300_000_000,300_000_000,7.2,date(2024,6,1),date(2025,6,1),365,"Vigente","002-001-001","Inversión Maquinaria","GAR002","","","",None,None,None],
            ["OP003","Empresa B","Banco Santander","Linea de Credito","LC-2024-001","CLP",200_000_000,120_000_000,5.8,date(2024,1,15),date(2025,1,15),365,"Vigente","003-001-001","Capital de Trabajo","GAR003","","","",None,None,None],
            ["OP004","Empresa A","Banco Chile","Refinanciamiento","REF-2024-001","CLP",250_000_000,250_000_000,6.0,date(2024,8,1),date(2026,8,1),730,"Vigente","001-001-002","Refinanc. OP001 parcial","GAR001","OP001","","",None,None,None],
            ["OP005","Empresa B","Banco BCI","Boleta de Garantia","BG-2024-050","CLP",50_000_000,50_000_000,0,date(2024,5,1),date(2025,5,1),365,"Vigente","","Licitación Pública","","","","",None,None,None],
        ]
    add_sheet("Base_Deuda", cols_bd, filas_bd)

    # ── Cronograma_Pagos ──
    cols_cp = [
        "ID_Pago","ID_Operacion","Empresa","Banco","Numero_Operacion_Banco",
        "Fecha_Pago","Capital","Interes","Comision","Seguro","Otros_Cargos",
        "Total_Pagar","Estado_Pago","Fecha_Pago_Real","Cuenta_Debito","Comprobante","Observaciones",
    ]
    filas_cp = []
    if con_ejemplo:
        filas_cp = [
            ["PAG001","OP001","Empresa A","Banco Chile","42356789",date(2025,6,1),50_000_000,2_437_500,0,0,0,52_437_500,"Pendiente",None,"001-001-001","","Cuota Jun-2025"],
            ["PAG002","OP001","Empresa A","Banco Chile","42356789",date(2025,7,1),50_000_000,2_166_667,0,0,0,52_166_667,"Pendiente",None,"001-001-001","","Cuota Jul-2025"],
            ["PAG003","OP001","Empresa A","Banco Chile","42356789",date(2025,8,1),50_000_000,1_895_833,0,0,0,51_895_833,"Pendiente",None,"001-001-001","","Cuota Ago-2025"],
            ["PAG004","OP001","Empresa A","Banco Chile","42356789",date(2025,9,1),300_000_000,1_625_000,0,0,0,301_625_000,"Pendiente",None,"001-001-001","","Pago final OP001"],
            ["PAG005","OP002","Empresa A","Banco BCI","78901234",date(2025,6,1),150_000_000,1_800_000,0,0,0,151_800_000,"Pendiente",None,"002-001-001","","Cuota semestral"],
            ["PAG006","OP002","Empresa A","Banco BCI","78901234",date(2025,12,1),150_000_000,900_000,0,0,0,150_900_000,"Pendiente",None,"002-001-001","","Pago final OP002"],
            ["PAG007","OP003","Empresa B","Banco Santander","LC-2024-001",date(2025,6,15),40_000_000,580_000,0,0,0,40_580_000,"Pendiente",None,"003-001-001","","Amortización línea"],
            ["PAG008","OP003","Empresa B","Banco Santander","LC-2024-001",date(2025,9,15),40_000_000,386_667,0,0,0,40_386_667,"Pendiente",None,"003-001-001","","Amortización línea"],
            ["PAG009","OP004","Empresa A","Banco Chile","REF-2024-001",date(2025,8,1),125_000_000,1_250_000,0,0,0,126_250_000,"Pendiente",None,"001-001-002","","Amortización semestral"],
            ["PAG010","OP004","Empresa A","Banco Chile","REF-2024-001",date(2026,2,1),125_000_000,625_000,0,0,0,125_625_000,"Pendiente",None,"001-001-002","","Pago final OP004"],
        ]
    add_sheet("Cronograma_Pagos", cols_cp, filas_cp)

    # ── Lineas_Credito ──
    cols_lc = [
        "ID_Linea","Empresa","Banco","Numero_Linea","Moneda",
        "Monto_Aprobado","Monto_Utilizado","Monto_Disponible","Tasa_Referencia",
        "Fecha_Aprobacion","Fecha_Vencimiento","Garantia","Estado","Observaciones",
    ]
    filas_lc = []
    if con_ejemplo:
        filas_lc = [
            ["LC001","Empresa A","Banco Chile","LCR-2024-001","CLP",500_000_000,350_000_000,150_000_000,5.5,date(2024,1,1),date(2026,1,1),"Hipoteca","Vigente","Línea revolving capital trabajo"],
            ["LC002","Empresa A","Banco BCI","LCR-BCI-2024","CLP",200_000_000,80_000_000,120_000_000,6.0,date(2024,3,1),date(2025,3,1),"Sin Garantia","Vigente","Línea para importaciones"],
            ["LC003","Empresa B","Banco Santander","LS-SAN-2024","USD",500_000,200_000,300_000,5.0,date(2024,6,1),date(2025,6,1),"Fianza","Vigente","Línea en dólares"],
        ]
    add_sheet("Lineas_Credito", cols_lc, filas_lc)

    # ── Boletas_Garantia ──
    cols_bg = [
        "ID_Boleta","Empresa","Banco","Numero_Boleta","Beneficiario","Moneda","Monto",
        "Fecha_Emision","Fecha_Vencimiento","Objeto","Contragarantia","Estado",
        "Dias_Al_Vencimiento","Operacion_Relacionada","Observaciones",
    ]
    filas_bg = []
    if con_ejemplo:
        filas_bg = [
            ["BG001","Empresa A","Banco Chile","BG-2024-1001","Ministerio de Obras Públicas","CLP",30_000_000,date(2024,3,1),date(2025,9,1),"Garantía licitación Proyecto Norte","Depósito a plazo","Vigente",None,"",""],
            ["BG002","Empresa A","Banco BCI","BG-2024-1002","Municipalidad de Santiago","CLP",15_000_000,date(2024,5,1),date(2025,6,30),"Garantía fiel cumplimiento contrato","Depósito a plazo","Vigente",None,"",""],
            ["BG003","Empresa B","Banco Santander","BG-2024-2001","CODELCO División Norte","USD",100_000,date(2024,6,1),date(2025,6,1),"Garantía anticipo contrato suministro","Warrant","Vigente",None,"OP005",""],
            ["BG004","Empresa B","Banco Chile","BG-2024-2002","Aguas del Norte S.A.","CLP",8_000_000,date(2024,2,1),date(2025,2,1),"Garantía seriedad oferta","Sin Garantia","Vencida",None,"","Pendiente renovación"],
            ["BG005","Empresa A","Banco Santander","BG-2024-1003","Metro de Santiago","CLP",25_000_000,date(2024,9,1),date(2025,9,1),"Garantía cumplimiento obra civil","Hipoteca","Vigente",None,"",""],
        ]
    add_sheet("Boletas_Garantia", cols_bg, filas_bg)

    # ── Refinanciamientos ──
    cols_rf = [
        "ID_Refinanciamiento","ID_Operacion_Anterior","ID_Operacion_Nueva","Empresa","Banco",
        "Monto_Cancelado","Monto_Nuevo","Fecha_Refinanciamiento","Motivo","Diferencia","Observaciones",
    ]
    filas_rf = []
    if con_ejemplo:
        filas_rf = [
            ["RF001","OP001","OP004","Empresa A","Banco Chile",250_000_000,250_000_000,date(2024,8,1),"Mejora de tasa y extensión de plazo",0,"OP001 refinanciada parcialmente; saldo restante continúa en OP001"],
        ]
    add_sheet("Refinanciamientos", cols_rf, filas_rf)

    # ── Bancos_Cuentas ──
    cols_bc = ["Empresa","Banco","Numero_Cuenta","Moneda","Tipo_Cuenta","Uso","Estado","Observaciones"]
    filas_bc = []
    if con_ejemplo:
        filas_bc = [
            ["Empresa A","Banco Chile","001-001-001-01","CLP","Corriente","Pago de deuda y operaciones","Activa",""],
            ["Empresa A","Banco BCI","002-001-001-01","CLP","Corriente","Pago deuda BCI","Activa",""],
            ["Empresa A","Banco Santander","003-001-001-01","CLP","Corriente","Pago deuda Santander","Activa",""],
            ["Empresa B","Banco Santander","003-001-001-02","USD","Corriente","Operaciones internacionales","Activa",""],
        ]
    add_sheet("Bancos_Cuentas", cols_bc, filas_bc)

    # ── Garantias ──
    cols_ga = [
        "ID_Garantia","Empresa","Banco","Tipo_Garantia","Descripcion_Garantia","Moneda",
        "Monto_Garantia","Operacion_Relacionada","Fecha_Inicio","Fecha_Vencimiento","Estado","Observaciones",
    ]
    filas_ga = []
    if con_ejemplo:
        filas_ga = [
            ["GAR001","Empresa A","Banco Chile","Hipoteca","Planta Industrial Norte – Lote 5","CLP",800_000_000,"OP001, OP004",date(2024,3,1),None,"Vigente","Hipoteca primer grado"],
            ["GAR002","Empresa A","Banco BCI","Prenda","Maquinaria CNC – Serie MX-2024","CLP",320_000_000,"OP002",date(2024,6,1),None,"Vigente","Prenda industrial sin desplazamiento"],
            ["GAR003","Empresa B","Banco Santander","Fianza","Fianza solidaria socios","CLP",250_000_000,"OP003",date(2024,1,15),None,"Vigente","Fianza personal socios mayoritarios"],
        ]
    add_sheet("Garantias", cols_ga, filas_ga)

    # ── Pagos_Realizados ──
    cols_pr = [
        "ID_Pago_Realizado","ID_Operacion","Empresa","Banco","Fecha_Pago",
        "Capital_Pagado","Interes_Pagado","Comision_Pagada","Total_Pagado",
        "Cuenta_Debito","Comprobante","Estado","Observaciones",
    ]
    add_sheet("Pagos_Realizados", cols_pr)

    # ── Parametros ──
    ws_p = wb.create_sheet("Parametros")
    ws_p.row_dimensions[1].height = 24
    for ci, h in enumerate(["Categoria","Valor"], 1):
        c = ws_p.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = brd
    param_rows = []
    for e in ["Empresa A","Empresa B","Empresa C"]:       param_rows.append(("Empresa", e))
    for b in ["Banco Chile","Banco BCI","Banco Santander","Banco Estado","Banco Itaú","Banco BBVA","Banco Security","Banco Scotiabank"]: param_rows.append(("Banco", b))
    for t in TIPOS_OPERACION:                             param_rows.append(("Tipo_Operacion", t))
    for m in MONEDAS:                                     param_rows.append(("Moneda", m))
    for e in ESTADOS_OPERACION:                           param_rows.append(("Estado_Operacion", e))
    for e in ESTADOS_PAGO:                                param_rows.append(("Estado_Pago", e))
    for e in ESTADOS_BOLETA:                              param_rows.append(("Estado_Boleta", e))
    for t in TIPOS_GARANTIA:                              param_rows.append(("Tipo_Garantia", t))
    for c in CLASIFICACIONES:                             param_rows.append(("Clasificacion_Vencimiento", c))
    for ri, (cat, val) in enumerate(param_rows, 2):
        ws_p.cell(ri, 1, cat).border = brd
        ws_p.cell(ri, 2, val).border = brd
    from openpyxl.utils import get_column_letter
    for ci in [1, 2]:
        mx = max(len(str(ws_p.cell(r, ci).value or "")) for r in range(1, len(param_rows)+2))
        ws_p.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 35)

    # ── Alertas (referencia) ──
    add_sheet("Alertas", ["Nivel","Mensaje","Referencia","Fecha_Deteccion","Responsable","Accion_Tomada"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# CARGA DE EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def cargar_excel(archivo) -> dict:
    """Carga todas las hojas disponibles desde un Excel."""
    hojas_esperadas = [
        "Base_Deuda","Cronograma_Pagos","Lineas_Credito","Boletas_Garantia",
        "Refinanciamientos","Bancos_Cuentas","Garantias","Pagos_Realizados",
        "Parametros","Alertas",
    ]
    dfs = {}
    try:
        xl = pd.ExcelFile(archivo)
        disponibles = set(xl.sheet_names)
        for hoja in hojas_esperadas:
            if hoja in disponibles:
                try:
                    dfs[hoja] = xl.parse(hoja)
                except Exception as e:
                    st.warning(f"No se pudo leer la hoja '{hoja}': {e}")
                    dfs[hoja] = pd.DataFrame()
            else:
                dfs[hoja] = pd.DataFrame()
        # Hojas extra
        for hoja in disponibles - set(hojas_esperadas):
            try:
                dfs[hoja] = xl.parse(hoja)
            except Exception:
                pass
    except Exception as e:
        st.error(f"Error al abrir el archivo: {e}")
        return {}
    return dfs


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR REPORTE
# ═══════════════════════════════════════════════════════════════════════════════

def exportar_reporte_excel(dfs: dict, kpis: dict, alertas: pd.DataFrame, trazabilidad: pd.DataFrame) -> bytes:
    """Genera reporte consolidado en Excel con formato profesional."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        wb  = writer.book
        fmt_hdr  = wb.add_format({"bold": True, "bg_color": "#1A3A5C", "font_color": "#FFFFFF", "border": 1, "align": "center"})
        fmt_num  = wb.add_format({"num_format": "#,##0", "border": 1})
        fmt_cell = wb.add_format({"border": 1})

        def write_df(df: pd.DataFrame, sheet: str):
            if df is None or df.empty:
                ws = wb.add_worksheet(sheet)
                ws.write(0, 0, "Sin datos disponibles", fmt_cell)
                return
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]
            for ci, col in enumerate(df.columns):
                ws.write(0, ci, col, fmt_hdr)
                ws.set_column(ci, ci, max(len(str(col)) + 2, 14))

        # Dashboard resumen
        ws_dash = wb.add_worksheet("Dashboard_Resumen")
        ws_dash.set_column(0, 0, 42)
        ws_dash.set_column(1, 1, 22)
        ws_dash.write(0, 0, "Indicador", fmt_hdr)
        ws_dash.write(0, 1, "Valor", fmt_hdr)
        resumen = [
            ("Deuda Bancaria Total",          kpis.get("deuda_total", 0)),
            ("Tasa Promedio Ponderada (%)",    kpis.get("tasa_promedio", 0)),
            ("Capital Próximos 7 Días",        kpis.get("capital_7d", 0)),
            ("Capital Próximos 30 Días",       kpis.get("capital_30d", 0)),
            ("Capital Próximos 60 Días",       kpis.get("capital_60d", 0)),
            ("Capital Próximos 90 Días",       kpis.get("capital_90d", 0)),
            ("Intereses Próximos 30 Días",     kpis.get("interes_30d", 0)),
            ("Intereses Próximos 90 Días",     kpis.get("interes_90d", 0)),
            ("Líneas de Crédito Aprobadas",    kpis.get("lineas_aprobadas", 0)),
            ("Líneas de Crédito Utilizadas",   kpis.get("lineas_utilizadas", 0)),
            ("Líneas de Crédito Disponibles",  kpis.get("lineas_disponibles", 0)),
            ("Boletas Vigentes (cantidad)",    kpis.get("boletas_vigentes", 0)),
            ("Boletas Vencidas (cantidad)",    kpis.get("boletas_vencidas", 0)),
            ("Operaciones Vencidas",           kpis.get("ops_vencidas", 0)),
            ("Operaciones Por Vencer 30d",     kpis.get("ops_por_vencer_30", 0)),
            ("Exposición Total Bancaria",      kpis.get("exposicion_total", 0)),
        ]
        for ri, (label, val) in enumerate(resumen, 1):
            ws_dash.write(ri, 0, label, fmt_cell)
            ws_dash.write(ri, 1, val, fmt_num)

        for hoja in ["Base_Deuda","Cronograma_Pagos","Lineas_Credito","Boletas_Garantia",
                     "Refinanciamientos","Garantias","Pagos_Realizados"]:
            write_df(dfs.get(hoja, pd.DataFrame()), hoja)

        write_df(alertas, "Alertas")
        write_df(trazabilidad, "Trazabilidad")

        df_deuda = dfs.get("Base_Deuda", pd.DataFrame())
        if not df_deuda.empty and "Dias_Al_Vencimiento" in df_deuda.columns:
            df_venc = df_deuda[df_deuda["Dias_Al_Vencimiento"].between(0, 90)].copy()
            write_df(df_venc, "Vencimientos")
        else:
            write_df(pd.DataFrame(), "Vencimientos")

    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS DE INTERFAZ
# ═══════════════════════════════════════════════════════════════════════════════

def fmt_m(val: float) -> str:
    """Formatea número grande de forma legible."""
    val = float(val or 0)
    if abs(val) >= 1_000_000_000:
        return f"${val/1_000_000_000:,.2f} MM"
    elif abs(val) >= 1_000_000:
        return f"${val/1_000_000:,.1f} M"
    else:
        return f"${val:,.0f}"


def color_nivel(nivel: str) -> str:
    return {"Crítica": "alert-critica", "Alta": "alert-alta",
            "Media": "alert-media", "Informativa": "alert-info"}.get(nivel, "alert-info")


def icono_nivel(nivel: str) -> str:
    return {"Crítica": "🔴", "Alta": "🟠", "Media": "🟡", "Informativa": "🔵"}.get(nivel, "•")


def tabla_filtrable(df: pd.DataFrame, key: str = "tbl"):
    if df is None or df.empty:
        st.info("Sin registros para mostrar.")
        return
    sel = st.multiselect("Columnas visibles", options=list(df.columns),
                          default=list(df.columns), key=f"{key}_cols")
    st.dataframe(df[sel] if sel else df, use_container_width=True, height=400)


def aplicar_filtros(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    for col, key in [("Empresa","f_empresa"),("Banco","f_banco"),("Moneda","f_moneda"),
                     ("Tipo_Operacion","f_tipo"),("Estado","f_estado")]:
        vals = st.session_state.get(key, [])
        if vals and col in df.columns:
            df = df[df[col].isin(vals)]
    f_desde = st.session_state.get("f_fecha_desde")
    f_hasta = st.session_state.get("f_fecha_hasta")
    if f_desde and f_hasta and "Fecha_Vencimiento" in df.columns:
        df = df[(df["Fecha_Vencimiento"] >= pd.Timestamp(f_desde)) &
                (df["Fecha_Vencimiento"] <= pd.Timestamp(f_hasta))]
    return df


def df_export_button(df: pd.DataFrame, label: str, filename: str, key: str):
    if df is None or df.empty:
        return
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
        df.to_excel(w, index=False, sheet_name="Datos")
    buf.seek(0)
    st.download_button(label=label, data=buf.read(), file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key=key)


# ═══════════════════════════════════════════════════════════════════════════════
# SECCIONES DE LA APLICACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

def seccion_inicio():
    st.markdown('<div class="main-header">🏦 Sistema de Control de Deuda Bancaria</div>', unsafe_allow_html=True)
    st.markdown("""
Herramienta diseñada para el equipo de **Tesorería** que centraliza el control, monitoreo y análisis de la deuda bancaria empresarial: préstamos directos, líneas de crédito, boletas de garantía, refinanciamientos y garantías.

---
### Módulos disponibles

| Módulo | Descripción |
|---|---|
| 📥 **Cargar Excel** | Suba su archivo Excel con la información bancaria |
| 📊 **Dashboard** | KPIs ejecutivos y gráficos de deuda |
| 💰 **Deuda Bancaria** | Tabla completa de operaciones de crédito |
| 📅 **Cronograma de Pagos** | Vencimientos de capital e intereses |
| 🏦 **Líneas de Crédito** | Uso y disponibilidad de líneas |
| 📋 **Boletas de Garantía** | Control de boletas vigentes y vencidas |
| 🔄 **Refinanciamientos** | Historial y trazabilidad de refinanciamientos |
| 🔒 **Garantías** | Registro de garantías asociadas |
| 💳 **Pagos Realizados** | Historial de pagos efectuados |
| 🔔 **Alertas** | Alertas críticas, altas, medias e informativas |
| 📤 **Exportar Reportes** | Reportes consolidados en Excel |

---
### Flujo de trabajo recomendado
1. Descargue la **plantilla Excel** desde **Cargar Excel**.
2. Complete la información en cada hoja.
3. Suba el archivo y revise el **Dashboard**.
4. Monitoree la sección de **Alertas** periódicamente.
5. Exporte reportes para distribución a Gerencia o Directorio.

> 💡 Puede usar la **plantilla con datos de ejemplo** para explorar el sistema sin datos reales.
""")


def seccion_cargar_excel():
    st.markdown('<div class="main-header">📥 Cargar Excel</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📄 Plantilla vacía")
        st.caption("Descargue la plantilla, complete la información y vuelva a cargarla.")
        data_vacia = crear_plantilla_excel(con_ejemplo=False)
        st.download_button("⬇️ Descargar plantilla vacía", data=data_vacia,
                           file_name="plantilla_deuda_bancaria.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col2:
        st.subheader("📊 Plantilla con datos de ejemplo")
        st.caption("Incluye 5 préstamos, 3 líneas, 5 boletas y 2 refinanciamientos de ejemplo.")
        data_ejemplo = crear_plantilla_excel(con_ejemplo=True)
        st.download_button("⬇️ Descargar datos de ejemplo", data=data_ejemplo,
                           file_name="plantilla_deuda_ejemplo.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("---")
    st.subheader("📂 Cargar archivo Excel")
    archivo = st.file_uploader("Seleccione su archivo (.xlsx)", type=["xlsx"])

    if archivo:
        with st.spinner("Procesando archivo..."):
            dfs = cargar_excel(archivo)
        if dfs:
            dfs = normalizar_datos(dfs)
            st.session_state["dfs"] = dfs
            st.success("✅ Archivo cargado y normalizado correctamente.")

            # Resumen de hojas cargadas
            st.subheader("Resumen de carga")
            info_hojas = [
                ("Base_Deuda","Operaciones"),("Cronograma_Pagos","Pagos prog."),
                ("Lineas_Credito","Líneas créd."),("Boletas_Garantia","Boletas"),
                ("Refinanciamientos","Refinanc."),("Garantias","Garantías"),
                ("Pagos_Realizados","Pagos real."),
            ]
            cols_res = st.columns(len(info_hojas))
            for i, (hoja, label) in enumerate(info_hojas):
                df = dfs.get(hoja, pd.DataFrame())
                cols_res[i].metric(label, len(df) if not df.empty else 0)

            # Vista rápida de alertas
            st.subheader("Resultado de validaciones")
            alertas = calcular_alertas(dfs)
            if alertas.empty:
                st.success("✅ Sin alertas detectadas.")
            else:
                crit = int((alertas["Nivel"] == "Crítica").sum())
                alta = int((alertas["Nivel"] == "Alta").sum())
                med  = int((alertas["Nivel"] == "Media").sum())
                info = int((alertas["Nivel"] == "Informativa").sum())
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("🔴 Críticas",    crit)
                c2.metric("🟠 Altas",       alta)
                c3.metric("🟡 Medias",      med)
                c4.metric("🔵 Informativas",info)
                if crit:
                    st.error(f"⚠️ Se detectaron {crit} alertas CRÍTICAS. Revise la sección **Alertas**.")
    elif "dfs" in st.session_state:
        st.info("ℹ️ Existe un archivo cargado en sesión. Suba uno nuevo para reemplazarlo.")


def seccion_dashboard():
    st.markdown('<div class="main-header">📊 Dashboard Ejecutivo</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})

    if not dfs:
        st.warning("Sin datos cargados. Use **Cargar Excel** o cargue los datos de ejemplo.")
        if st.button("▶ Cargar datos de ejemplo ahora"):
            raw = crear_plantilla_excel(con_ejemplo=True)
            dfs_ej = cargar_excel(io.BytesIO(raw))
            dfs_ej = normalizar_datos(dfs_ej)
            st.session_state["dfs"] = dfs_ej
            st.rerun()
        return

    kpis    = calcular_kpis(dfs)
    alertas = calcular_alertas(dfs)
    df_deuda= dfs.get("Base_Deuda", pd.DataFrame())
    df_cron = dfs.get("Cronograma_Pagos", pd.DataFrame())
    df_lineas= dfs.get("Lineas_Credito", pd.DataFrame())
    df_boletas= dfs.get("Boletas_Garantia", pd.DataFrame())

    # ── KPIs fila 1 ──
    st.subheader("Indicadores Clave")
    r1 = st.columns(5)
    r1[0].metric("💰 Deuda Total",          fmt_m(kpis["deuda_total"]))
    r1[1].metric("📈 Tasa Prom. Ponderada", f"{kpis['tasa_promedio']:.2f}%")
    r1[2].metric("📅 Capital 30d",          fmt_m(kpis["capital_30d"]))
    r1[3].metric("💵 Intereses 30d",        fmt_m(kpis["interes_30d"]))
    r1[4].metric("🌐 Exposición Total",     fmt_m(kpis["exposicion_total"]))

    r2 = st.columns(5)
    r2[0].metric("✅ Líneas Aprobadas",     fmt_m(kpis["lineas_aprobadas"]))
    r2[1].metric("📊 Líneas Utilizadas",    fmt_m(kpis["lineas_utilizadas"]))
    r2[2].metric("🟢 Líneas Disponibles",   fmt_m(kpis["lineas_disponibles"]))
    r2[3].metric("📋 Boletas Vigentes",     kpis["boletas_vigentes"])
    r2[4].metric("⚠️ Ops Vencidas",         kpis["ops_vencidas"])

    st.markdown("---")

    # ── Fila gráficos 1 ──
    g1, g2 = st.columns(2)
    with g1:
        st.subheader("Deuda por Banco")
        if not kpis["deuda_por_banco"].empty:
            fig = px.bar(kpis["deuda_por_banco"].sort_values("Saldo_Actual"),
                         x="Saldo_Actual", y="Banco", orientation="h",
                         color="Saldo_Actual", color_continuous_scale="Blues",
                         labels={"Saldo_Actual": "Saldo", "Banco": ""})
            fig.update_layout(showlegend=False, height=300, margin=dict(l=0,r=10,t=10,b=0))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin datos.")

    with g2:
        st.subheader("Deuda por Tipo de Operación")
        if not kpis["deuda_por_tipo"].empty:
            fig = px.pie(kpis["deuda_por_tipo"], values="Saldo_Actual", names="Tipo_Operacion",
                         color_discrete_sequence=px.colors.sequential.Blues_r)
            fig.update_layout(height=300, margin=dict(l=0,r=0,t=10,b=0))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin datos.")

    # ── Fila gráficos 2 ──
    g3, g4 = st.columns(2)
    with g3:
        st.subheader("Deuda por Empresa")
        if not kpis["deuda_por_empresa"].empty:
            fig = px.bar(kpis["deuda_por_empresa"], x="Empresa", y="Saldo_Actual",
                         color="Empresa", color_discrete_sequence=px.colors.sequential.Blues_r,
                         labels={"Saldo_Actual": "Saldo", "Empresa": ""})
            fig.update_layout(showlegend=False, height=280, margin=dict(l=0,r=0,t=10,b=0))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin datos.")

    with g4:
        st.subheader("Deuda por Moneda")
        if not kpis["deuda_por_moneda"].empty:
            fig = px.pie(kpis["deuda_por_moneda"], values="Saldo_Actual", names="Moneda",
                         color_discrete_sequence=["#1A3A5C","#2E6DA4","#5B9BD5","#A9C6E3"])
            fig.update_layout(height=280, margin=dict(l=0,r=0,t=10,b=0))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin datos.")

    # ── Vencimientos por mes ──
    st.subheader("Capital e Intereses por Mes (próximos 12 meses)")
    today_ts = pd.Timestamp(HOY)
    if not df_cron.empty and "Fecha_Pago" in df_cron.columns and "Capital" in df_cron.columns:
        futuro = df_cron[(df_cron["Fecha_Pago"] >= today_ts) &
                         (df_cron["Fecha_Pago"] <= today_ts + pd.DateOffset(months=12))].copy()
        if not futuro.empty:
            futuro["Mes"] = futuro["Fecha_Pago"].dt.to_period("M").astype(str)
            agr = futuro.groupby("Mes").agg(Capital=("Capital","sum"),
                                             Interes=("Interes","sum")).reset_index()
            fig = go.Figure()
            fig.add_bar(x=agr["Mes"], y=agr["Capital"], name="Capital",   marker_color="#1A3A5C")
            fig.add_bar(x=agr["Mes"], y=agr["Interes"], name="Intereses", marker_color="#5B9BD5")
            fig.update_layout(barmode="stack", height=300,
                              margin=dict(l=0,r=0,t=10,b=0),
                              legend=dict(orientation="h", yanchor="bottom", y=1.02))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin vencimientos futuros en el cronograma.")
    else:
        st.info("Sin datos de cronograma de pagos.")

    # ── Fila gráficos 3 ──
    g5, g6 = st.columns(2)
    with g5:
        st.subheader("Líneas de Crédito: Aprobado vs Utilizado")
        if not df_lineas.empty and "Banco" in df_lineas.columns:
            fig = go.Figure()
            fig.add_bar(x=df_lineas["Banco"], y=df_lineas.get("Monto_Aprobado",  pd.Series()), name="Aprobado",  marker_color="#A9C6E3")
            fig.add_bar(x=df_lineas["Banco"], y=df_lineas.get("Monto_Utilizado", pd.Series()), name="Utilizado", marker_color="#1A3A5C")
            fig.update_layout(barmode="overlay", height=290,
                              margin=dict(l=0,r=0,t=10,b=0),
                              legend=dict(orientation="h", yanchor="bottom", y=1.02))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin datos de líneas de crédito.")

    with g6:
        st.subheader("Boletas por Estado")
        if not df_boletas.empty and "Estado" in df_boletas.columns and "Monto" in df_boletas.columns:
            agr = df_boletas.groupby("Estado")["Monto"].sum().reset_index()
            fig = px.bar(agr, x="Estado", y="Monto", color="Estado",
                         color_discrete_sequence=px.colors.sequential.Blues_r)
            fig.update_layout(showlegend=False, height=290, margin=dict(l=0,r=0,t=10,b=0))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Sin datos de boletas.")

    # ── Ranking exposición ──
    st.subheader("Ranking de Exposición Bancaria Total")
    df_exp = calcular_exposicion_por_banco(dfs)
    if not df_exp.empty:
        fig = px.bar(df_exp.sort_values("Exposicion_Total"),
                     x="Exposicion_Total", y="Banco", orientation="h",
                     color="Exposicion_Total", color_continuous_scale="Blues",
                     labels={"Exposicion_Total": "Exposición Total", "Banco": ""})
        fig.update_layout(showlegend=False, height=280, margin=dict(l=0,r=10,t=10,b=0))
        st.plotly_chart(fig, use_container_width=True)

    # ── Alertas críticas rápidas ──
    criticas = alertas[alertas["Nivel"] == "Crítica"] if not alertas.empty else pd.DataFrame()
    if not criticas.empty:
        st.markdown("---")
        st.subheader("🔴 Alertas Críticas")
        for _, row in criticas.iterrows():
            st.markdown(f'<div class="alert-critica">🔴 <strong>{row["Mensaje"]}</strong> — <em>{row["Referencia"]}</em></div>',
                        unsafe_allow_html=True)


def seccion_deuda_bancaria():
    st.markdown('<div class="main-header">💰 Deuda Bancaria</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    df  = dfs.get("Base_Deuda", pd.DataFrame())
    if df.empty:
        st.info("Sin datos de deuda bancaria cargados.")
        return

    df_f = aplicar_filtros(df)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Operaciones",   len(df_f))
    c2.metric("Saldo total",   fmt_m(df_f["Saldo_Actual"].sum() if "Saldo_Actual" in df_f.columns else 0))
    c3.metric("Vencidas",      int((df_f["Dias_Al_Vencimiento"] < 0).sum()) if "Dias_Al_Vencimiento" in df_f.columns else 0)
    c4.metric("Interés mens.", fmt_m(df_f["Interes_Estimado_Mensual"].sum() if "Interes_Estimado_Mensual" in df_f.columns else 0))

    st.markdown("---")
    tab1, tab2, tab3 = st.tabs(["📋 Tabla completa", "⏰ Próximas a vencer (90d)", "❌ Vencidas"])
    with tab1:
        tabla_filtrable(df_f, "bd_full")
        df_export_button(df_f, "⬇️ Exportar tabla", f"deuda_{HOY.strftime('%Y%m%d')}.xlsx", "exp_bd")
    with tab2:
        if "Dias_Al_Vencimiento" in df_f.columns:
            sub = df_f[df_f["Dias_Al_Vencimiento"].between(0, 90)].sort_values("Dias_Al_Vencimiento")
            tabla_filtrable(sub, "bd_prox")
        else:
            st.info("Sin info de vencimientos.")
    with tab3:
        if "Dias_Al_Vencimiento" in df_f.columns and "Estado" in df_f.columns:
            sub = df_f[(df_f["Dias_Al_Vencimiento"] < 0) & (df_f["Estado"] == "Vigente")]
            tabla_filtrable(sub, "bd_venc")
        else:
            st.info("Sin info.")


def seccion_cronograma():
    st.markdown('<div class="main-header">📅 Cronograma de Pagos</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    df  = dfs.get("Cronograma_Pagos", pd.DataFrame())
    if df.empty:
        st.info("Sin datos de cronograma cargados.")
        return

    today_ts = pd.Timestamp(HOY)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total cuotas", len(df))
    if "Fecha_Pago" in df.columns and "Capital" in df.columns:
        m30 = (df["Fecha_Pago"] >= today_ts) & (df["Fecha_Pago"] <= today_ts + pd.Timedelta(days=30))
        c2.metric("Capital 30d",   fmt_m(df[m30]["Capital"].sum()))
        c3.metric("Intereses 30d", fmt_m(df[m30]["Interes"].sum() if "Interes" in df.columns else 0))
        c4.metric("Total 30d",     fmt_m(df[m30]["Total_Pagar"].sum() if "Total_Pagar" in df.columns else 0))

    st.markdown("---")
    tab1, tab2 = st.tabs(["📋 Cronograma completo", "⏰ Próximos 90 días"])
    with tab1:
        tabla_filtrable(df, "cp_full")
        df_export_button(df, "⬇️ Exportar cronograma", f"cronograma_{HOY.strftime('%Y%m%d')}.xlsx", "exp_cp")
    with tab2:
        if "Fecha_Pago" in df.columns:
            prox = df[(df["Fecha_Pago"] >= today_ts) &
                      (df["Fecha_Pago"] <= today_ts + pd.Timedelta(days=90))].sort_values("Fecha_Pago")
            tabla_filtrable(prox, "cp_90")
        else:
            st.info("Sin información de fechas.")


def seccion_lineas_credito():
    st.markdown('<div class="main-header">🏦 Líneas de Crédito</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    df  = dfs.get("Lineas_Credito", pd.DataFrame())
    if df.empty:
        st.info("Sin datos de líneas de crédito cargados.")
        return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total líneas",    len(df))
    c2.metric("Total aprobado",  fmt_m(df["Monto_Aprobado"].sum()  if "Monto_Aprobado"  in df.columns else 0))
    c3.metric("Total utilizado", fmt_m(df["Monto_Utilizado"].sum() if "Monto_Utilizado" in df.columns else 0))
    c4.metric("Total disponible",fmt_m(df["Monto_Disponible"].sum()if "Monto_Disponible"in df.columns else 0))

    st.markdown("---")
    tabla_filtrable(df, "lc")
    df_export_button(df, "⬇️ Exportar líneas", f"lineas_{HOY.strftime('%Y%m%d')}.xlsx", "exp_lc")


def seccion_boletas():
    st.markdown('<div class="main-header">📋 Boletas de Garantía</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    df  = dfs.get("Boletas_Garantia", pd.DataFrame())
    if df.empty:
        st.info("Sin datos de boletas de garantía cargados.")
        return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total boletas", len(df))
    c2.metric("Vigentes",  int((df["Estado"] == "Vigente").sum()) if "Estado" in df.columns else 0)
    c3.metric("Vencidas",  int((df["Estado"] == "Vencida").sum()) if "Estado" in df.columns else 0)
    c4.metric("Monto total",fmt_m(df["Monto"].sum() if "Monto" in df.columns else 0))

    st.markdown("---")
    tab1, tab2, tab3 = st.tabs(["📋 Todas", "✅ Vigentes", "⚠️ Vencidas / Por vencer"])
    with tab1:
        tabla_filtrable(df, "bg_all")
        df_export_button(df, "⬇️ Exportar boletas", f"boletas_{HOY.strftime('%Y%m%d')}.xlsx", "exp_bg")
    with tab2:
        sub = df[df["Estado"] == "Vigente"] if "Estado" in df.columns else df
        tabla_filtrable(sub, "bg_vig")
    with tab3:
        if "Estado" in df.columns and "Dias_Al_Vencimiento" in df.columns:
            por_vencer = df[(df["Estado"] == "Vigente") &
                            (df["Dias_Al_Vencimiento"] >= 0) &
                            (df["Dias_Al_Vencimiento"] <= 30)]
            vencidas   = df[df["Estado"] == "Vencida"]
            st.caption("**Por vencer dentro de 30 días**")
            tabla_filtrable(por_vencer, "bg_pv")
            st.caption("**Boletas vencidas**")
            tabla_filtrable(vencidas, "bg_ve")
        else:
            st.info("Sin información de fechas / estado.")


def seccion_refinanciamientos():
    st.markdown('<div class="main-header">🔄 Refinanciamientos y Trazabilidad</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})

    tab1, tab2 = st.tabs(["📋 Refinanciamientos", "🔗 Trazabilidad de operaciones"])
    with tab1:
        df = dfs.get("Refinanciamientos", pd.DataFrame())
        if df.empty:
            st.info("Sin datos de refinanciamientos cargados.")
        else:
            tabla_filtrable(df, "rf")
            df_export_button(df, "⬇️ Exportar refinanciamientos",
                             f"refinanciamientos_{HOY.strftime('%Y%m%d')}.xlsx", "exp_rf")

    with tab2:
        st.caption("""
        Muestra la relación entre operaciones anteriores y nuevas generadas por renovaciones o refinanciamientos.
        El flujo esperado es: **Préstamo original → amortización → renovación/refinanciamiento → nuevo préstamo → pago/cierre**.
        """)
        traz = calcular_trazabilidad(dfs)
        if traz.empty:
            st.info("Sin información de trazabilidad. Complete la hoja **Refinanciamientos** o el campo **ID_Operacion_Reemplazada** en Base_Deuda.")
        else:
            tabla_filtrable(traz, "traz")
            df_export_button(traz, "⬇️ Exportar trazabilidad",
                             f"trazabilidad_{HOY.strftime('%Y%m%d')}.xlsx", "exp_traz")


def seccion_garantias():
    st.markdown('<div class="main-header">🔒 Garantías</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    df  = dfs.get("Garantias", pd.DataFrame())
    if df.empty:
        st.info("Sin datos de garantías cargados.")
        return

    c1, c2 = st.columns(2)
    c1.metric("Total garantías",  len(df))
    c2.metric("Monto total",      fmt_m(df["Monto_Garantia"].sum() if "Monto_Garantia" in df.columns else 0))

    st.markdown("---")
    tabla_filtrable(df, "gar")
    df_export_button(df, "⬇️ Exportar garantías", f"garantias_{HOY.strftime('%Y%m%d')}.xlsx", "exp_gar")


def seccion_pagos_realizados():
    st.markdown('<div class="main-header">💳 Pagos Realizados</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    df  = dfs.get("Pagos_Realizados", pd.DataFrame())
    if df.empty:
        st.info("Sin datos de pagos realizados cargados.")
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Total pagos",     len(df))
    c2.metric("Capital pagado",  fmt_m(df["Capital_Pagado"].sum()  if "Capital_Pagado"  in df.columns else 0))
    c3.metric("Total pagado",    fmt_m(df["Total_Pagado"].sum()    if "Total_Pagado"    in df.columns else 0))

    st.markdown("---")
    tabla_filtrable(df, "pr")
    df_export_button(df, "⬇️ Exportar pagos", f"pagos_realizados_{HOY.strftime('%Y%m%d')}.xlsx", "exp_pr")


def seccion_alertas():
    st.markdown('<div class="main-header">🔔 Alertas del Sistema</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    if not dfs:
        st.info("Sin datos cargados. Vaya a **Cargar Excel** primero.")
        return

    alertas = calcular_alertas(dfs)
    if alertas.empty:
        st.success("✅ Sin alertas detectadas. La información está en orden.")
        return

    # Resumen por nivel
    crit = int((alertas["Nivel"] == "Crítica").sum())
    alta = int((alertas["Nivel"] == "Alta").sum())
    med  = int((alertas["Nivel"] == "Media").sum())
    info = int((alertas["Nivel"] == "Informativa").sum())
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🔴 Críticas",    crit)
    c2.metric("🟠 Altas",       alta)
    c3.metric("🟡 Medias",      med)
    c4.metric("🔵 Informativas",info)

    st.markdown("---")
    niveles_sel = st.multiselect("Filtrar por nivel de alerta",
                                  options=["Crítica","Alta","Media","Informativa"],
                                  default=["Crítica","Alta","Media","Informativa"])
    filtradas = alertas[alertas["Nivel"].isin(niveles_sel)]

    for _, row in filtradas.iterrows():
        cls   = color_nivel(row["Nivel"])
        icono = icono_nivel(row["Nivel"])
        st.markdown(
            f'<div class="{cls}">{icono} <strong>[{row["Nivel"]}]</strong> {row["Mensaje"]}'
            f'&nbsp; | &nbsp;<em>{row["Referencia"]}</em></div>',
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.subheader("Tabla de alertas")
    st.dataframe(filtradas, use_container_width=True)
    df_export_button(filtradas, "⬇️ Exportar alertas",
                     f"alertas_{HOY.strftime('%Y%m%d')}.xlsx", "exp_alertas")


def seccion_exportar():
    st.markdown('<div class="main-header">📤 Exportar Reportes</div>', unsafe_allow_html=True)
    dfs = st.session_state.get("dfs", {})
    if not dfs:
        st.info("Sin datos cargados. Vaya a **Cargar Excel** primero.")
        return

    kpis         = calcular_kpis(dfs)
    alertas      = calcular_alertas(dfs)
    trazabilidad = calcular_trazabilidad(dfs)

    st.subheader("Reportes disponibles")
    col1, col2 = st.columns(2)

    with col1:
        # Reporte consolidado
        reporte = exportar_reporte_excel(dfs, kpis, alertas, trazabilidad)
        st.download_button(
            label="📊 Reporte Consolidado (todas las hojas)",
            data=reporte,
            file_name=f"reporte_consolidado_{HOY.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_consolidado",
        )
        st.caption("Incluye Dashboard_Resumen, Base_Deuda, Cronograma, Líneas, Boletas, Refinanciamientos, Garantías, Pagos, Alertas, Trazabilidad, Vencimientos.")

        st.markdown("---")
        df_export_button(alertas, "🔔 Reporte de Alertas",
                         f"alertas_{HOY.strftime('%Y%m%d')}.xlsx", "dl_alertas")

        df_deuda = dfs.get("Base_Deuda", pd.DataFrame())
        if not df_deuda.empty and "Dias_Al_Vencimiento" in df_deuda.columns:
            venc90 = df_deuda[df_deuda["Dias_Al_Vencimiento"].between(0, 90)].copy()
            df_export_button(venc90, "📅 Vencimientos 90 días",
                             f"vencimientos_{HOY.strftime('%Y%m%d')}.xlsx", "dl_venc")

    with col2:
        df_export_button(dfs.get("Boletas_Garantia", pd.DataFrame()),
                         "📋 Boletas de Garantía",
                         f"boletas_{HOY.strftime('%Y%m%d')}.xlsx", "dl_boletas")

        df_export_button(dfs.get("Lineas_Credito", pd.DataFrame()),
                         "🏦 Líneas de Crédito",
                         f"lineas_{HOY.strftime('%Y%m%d')}.xlsx", "dl_lineas")

        df_export_button(trazabilidad,
                         "🔄 Trazabilidad de Refinanciamientos",
                         f"trazabilidad_{HOY.strftime('%Y%m%d')}.xlsx", "dl_traz")

        df_exp = calcular_exposicion_por_banco(dfs)
        df_export_button(df_exp,
                         "🏛️ Exposición por Banco",
                         f"exposicion_{HOY.strftime('%Y%m%d')}.xlsx", "dl_exp")


# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR Y NAVEGACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

def renderizar_sidebar() -> str:
    with st.sidebar:
        st.markdown("## 🏦 Deuda Bancaria")
        st.markdown("---")

        seccion = st.radio(
            "Menú",
            options=[
                "🏠 Inicio",
                "📥 Cargar Excel",
                "📊 Dashboard",
                "💰 Deuda Bancaria",
                "📅 Cronograma de Pagos",
                "🏦 Líneas de Crédito",
                "📋 Boletas de Garantía",
                "🔄 Refinanciamientos",
                "🔒 Garantías",
                "💳 Pagos Realizados",
                "🔔 Alertas",
                "📤 Exportar Reportes",
            ],
            label_visibility="collapsed",
        )

        st.markdown("---")
        # Estado de sesión
        if "dfs" in st.session_state:
            dfs    = st.session_state["dfs"]
            df_bd  = dfs.get("Base_Deuda", pd.DataFrame())
            n_ops  = len(df_bd) if not df_bd.empty else 0
            alertas= calcular_alertas(dfs)
            n_crit = int((alertas["Nivel"] == "Crítica").sum()) if not alertas.empty else 0
            st.success(f"✅ Datos cargados\n**{n_ops}** operaciones")
            if n_crit:
                st.error(f"🔴 {n_crit} alerta(s) crítica(s)")
            if st.button("🗑️ Limpiar sesión"):
                del st.session_state["dfs"]
                st.rerun()
        else:
            st.warning("⚠️ Sin datos cargados")

        # Filtros globales
        if "dfs" in st.session_state:
            df_deuda = st.session_state["dfs"].get("Base_Deuda", pd.DataFrame())
            if not df_deuda.empty:
                st.markdown("---")
                st.markdown("**Filtros globales**")
                for col, key, lbl in [
                    ("Empresa",       "f_empresa", "Empresa"),
                    ("Banco",         "f_banco",   "Banco"),
                    ("Moneda",        "f_moneda",  "Moneda"),
                    ("Tipo_Operacion","f_tipo",    "Tipo Operación"),
                    ("Estado",        "f_estado",  "Estado"),
                ]:
                    if col in df_deuda.columns:
                        opts = sorted(df_deuda[col].dropna().unique().tolist())
                        st.multiselect(lbl, options=opts, default=[], key=key)

                if "Fecha_Vencimiento" in df_deuda.columns:
                    fv = df_deuda["Fecha_Vencimiento"].dropna()
                    if not fv.empty:
                        st.date_input("Vencimiento desde", value=fv.min().date(), key="f_fecha_desde")
                        st.date_input("Vencimiento hasta", value=fv.max().date(), key="f_fecha_hasta")

    return seccion


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    seccion = renderizar_sidebar()

    if   seccion == "🏠 Inicio":               seccion_inicio()
    elif seccion == "📥 Cargar Excel":          seccion_cargar_excel()
    elif seccion == "📊 Dashboard":             seccion_dashboard()
    elif seccion == "💰 Deuda Bancaria":        seccion_deuda_bancaria()
    elif seccion == "📅 Cronograma de Pagos":   seccion_cronograma()
    elif seccion == "🏦 Líneas de Crédito":     seccion_lineas_credito()
    elif seccion == "📋 Boletas de Garantía":   seccion_boletas()
    elif seccion == "🔄 Refinanciamientos":     seccion_refinanciamientos()
    elif seccion == "🔒 Garantías":             seccion_garantias()
    elif seccion == "💳 Pagos Realizados":      seccion_pagos_realizados()
    elif seccion == "🔔 Alertas":               seccion_alertas()
    elif seccion == "📤 Exportar Reportes":     seccion_exportar()


if __name__ == "__main__":
    main()
